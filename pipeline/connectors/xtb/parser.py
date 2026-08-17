"""XTB new-format Excel (.xlsx) report parsing.

Parses the new XTB 3-sheet report ("Open Positions", "Cash Operations",
"Closed Positions") into typed dataclasses. Uses openpyxl (D13) to load
the workbook; date-formatted cells are auto-converted to naive ``datetime``
and the parser attaches ``tzinfo=UTC`` at the boundary (D3).

The parser is the single source of truth for:
  - account currency (Open Positions summary block Currency, D5);
  - per-ticker aggregate holdings (child lot rows skipped, D4);
  - the full cash ledger (Cash Operations, with Total row read into
    ``XtbReport.free_cash`` before being excluded from events, and
    ``Subaccount transfer`` rows filtered out, D7/D10/D22);
  - Closed Positions fee-enrichment fields keyed by Position ID (D8).

Guards 1, 3, 4, 5 are implemented here; guards 2, 6, 7, 8, 9, 10 are
transform-side (see the overhaul plan §3 Stage 1b) and are handled by the
transform, not this module.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel
from openpyxl.workbook.workbook import Workbook


class XtbError(RuntimeError):
    """Raised when an XTB report cannot be parsed (e.g. missing account currency)."""


# ---------------------------------------------------------------------------
# Dataclasses (binding field scope per the overhaul plan — YAGNI).
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class XtbOpenPosition:
    """A per-ticker aggregate open-position row (child lots are skipped, D4)."""

    account_id: str
    product: str  # "Investment Plan" | "My Trades" group label; not mapped downstream
    instrument: str  # real instrument name on the aggregate row -> description
    ticker: str  # reliable identity key -> label / identifier
    category: (
        str  # e.g. "ETF" on the aggregate row -> asset_class (empty on child lots)
    )
    value: (
        float  # account-currency market value (aggregate total), 2dp -> security_value
    )


@dataclass(frozen=True)
class XtbClosedPosition:
    """Fee-enrichment lookup row (D2); never emitted as its own CDC event."""

    position_id: str  # join key to Cash Operations trade rows
    commission: float  # -> fee_amount on the closing (Stock sell) row only (D8)
    purchase_value: float  # -> gross_amount = sale_value - purchase_value (transform)
    sale_value: float
    close_time: datetime  # -> settle_date on the closing row (UTC)


@dataclass(frozen=True)
class XtbCashOperation:
    """A single Cash Operations row (Total/summary rows excluded; subaccount
    transfers filtered out by the parser, D7)."""

    account_id: str
    operation_type: str  # raw "Type" text -> raw_event_type / event_type
    ticker: str  # populated on trade rows; empty on non-trade rows
    time: datetime  # UTC -> event_datetime
    amount: float  # account currency, 2dp -> cash_amount
    operation_id: str  # -> event_id (CDC dedup key)
    comment: str  # -> description; carries trade qty/price + transfer FX details
    position_id: str  # join key to Closed Positions (trade rows only)


@dataclass(frozen=True)
class XtbReport:
    """Parsed XTB report: all 3 sheets' extracted rows plus account metadata."""

    account_id: str
    account_ccy: str  # summary-block Currency (D5)
    open_positions: list[
        XtbOpenPosition
    ]  # per-ticker aggregate rows (child lots skipped)
    closed_positions: list[XtbClosedPosition]
    cash_operations: list[XtbCashOperation]  # Total/summary rows excluded from events
    free_cash: (
        float | None
    )  # Cash Operations Total -> snapshot CASH holding (D22); None if absent


# ---------------------------------------------------------------------------
# Helpers (kept from the old parser where still useful).
# ---------------------------------------------------------------------------


def normalize_header(value: Any) -> str:
    """Normalize a header cell to a lowercase single-space string for matching."""
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def find_sheet_name(sheet_names: list[str], expected_fragment: str) -> str:
    """Return the sheet name containing ``expected_fragment`` (case-insensitive).

    Raises ``XtbError`` if no sheet matches. Used for the 3 canonical sheets;
    callers wrap it to implement guard 5 (missing sheet -> empty list).
    """
    fragment = expected_fragment.lower()
    for name in sheet_names:
        if fragment in name.lower():
            return name
    raise XtbError(f"Could not find XTB sheet containing '{expected_fragment}'.")


def _str_cell(value: Any) -> str:
    """Coerce a cell value to a stripped string, mapping ``None`` to ``""``.

    Guard 1: openpyxl may return ``int`` for a numeric Position ID on one
    sheet and ``str`` on the other; without coercion the join silently
    misses. ``str(None).strip()`` would yield ``"None"``, so ``None`` is
    mapped to the empty string before ``str`` is applied.
    """
    if value is None:
        return ""
    return str(value).strip()


def as_float(value: Any, default: float = 0.0) -> float:
    """Best-effort float coercion for numeric cells (None/"" -> default)."""
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    normalized = str(value).strip().replace("\xa0", "").replace(" ", "")
    if "," in normalized and "." not in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return float(normalized)
    except (TypeError, ValueError):
        return default


def _to_utc_datetime(value: Any) -> datetime | None:
    """Decode a date cell to a timezone-aware UTC datetime (D3).

    openpyxl auto-converts date-formatted cells to naive ``datetime``; attach
    ``tzinfo=UTC`` at the parser boundary. For a raw numeric serial
    (defensive branch, not observed in the sample), pass it through
    ``openpyxl.utils.datetime.from_excel`` (handles the 1900 epoch and the
    1900-02-29 leap-year bug) then attach UTC. Returns ``None`` for empty
    cells so callers can skip malformed rows.
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo is not None else value.replace(tzinfo=UTC)
    if isinstance(value, (int, float)):
        decoded = from_excel(value)
        return decoded if decoded.tzinfo is not None else decoded.replace(tzinfo=UTC)
    raise XtbError(f"Unparseable XTB date cell: {value!r}")


def _row_cells(row: tuple[Any, ...]) -> list[Any]:
    """Normalize a row tuple to a list (openpyxl returns tuples)."""
    return list(row)


# ---------------------------------------------------------------------------
# Sheet loading.
# ---------------------------------------------------------------------------


def _load_workbook(data: bytes) -> Workbook:
    """Load an .xlsx byte string with openpyxl (data_only=True, D13)."""
    try:
        return openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:  # openpyxl raises a variety of error types
        raise XtbError("XTB report bytes are not a valid .xlsx file") from exc


def _sheet_rows(wb: Workbook, fragment: str) -> list[list[Any]] | None:
    """Return the rows of the sheet whose name contains ``fragment``.

    Returns ``None`` when the sheet is absent (guard 5: missing sheet does
    not abort the whole parse — the caller returns an empty list).
    """
    try:
        name = find_sheet_name(list(wb.sheetnames), fragment)
    except XtbError:
        return None
    ws = wb[name]
    return [_row_cells(row) for row in ws.iter_rows(values_only=True)]


def _account_id_from_rows(rows: list[list[Any]] | None) -> str | None:
    """Read the account id from R1 ``Account number`` (col 2) of a sheet."""
    if not rows:
        return None
    first = rows[0]
    if len(first) >= 2 and normalize_header(first[0]) == "account number":
        return _str_cell(first[1]) or None
    return None


# ---------------------------------------------------------------------------
# Open Positions (D4/D5, guards 3 and 4).
# ---------------------------------------------------------------------------

# Detail-header signature: row[0]=="Product" and row[1] starts with "Instrument".
_OP_DETAIL_HEADER_COL1 = ("instrument/position", "instrument")


def _parse_open_positions(
    rows: list[list[Any]],
    account_id: str,
) -> tuple[list[XtbOpenPosition], str]:
    """Parse Open Positions -> (aggregate holdings, account_ccy).

    The account currency is read from the summary block's Currency column
    (D5); guard 4 raises ``XtbError`` if it is empty or the block is absent.
    Per-ticker aggregate rows are kept (empty ``Type``, real name in
    ``Instrument``, non-empty ``Category``); child lot rows (non-empty
    ``Type``) are skipped (D4). Zero-value aggregates are skipped (guard 3).
    """
    account_ccy = _read_summary_block_currency(rows)
    if not account_ccy:
        raise XtbError("account currency missing from summary block")

    header_index = _find_detail_header(rows, _OP_DETAIL_HEADER_COL1)
    if header_index is None:
        return [], account_ccy

    positions: list[XtbOpenPosition] = []
    for row in rows[header_index + 1 :]:
        if not row or row[0] is None:
            continue
        # Column layout (1-based in the sheet, 0-based here):
        # 0 Product | 1 Instrument/Position | 2 Ticker | 3 Category |
        # 4 Type | 5 Volume | 6 Value | ...
        type_cell = _str_cell(row[4]) if len(row) > 4 else ""
        # Aggregate row: empty Type. Child lot row: non-empty Type -> skip (D4).
        if type_cell:
            continue
        ticker = _str_cell(row[2]) if len(row) > 2 else ""
        if not ticker:
            continue
        value = round(as_float(row[6]) if len(row) > 6 else 0.0, 2)
        # Guard 3: skip zero-value aggregate rows (delisted/fully-sold).
        if value == 0:
            continue
        positions.append(
            XtbOpenPosition(
                account_id=account_id,
                product=_str_cell(row[0]) if len(row) > 0 else "",
                instrument=_str_cell(row[1]) if len(row) > 1 else "",
                ticker=ticker,
                category=_str_cell(row[3]) if len(row) > 3 else "",
                value=value,
            )
        )
    return positions, account_ccy


def _read_summary_block_currency(rows: list[list[Any]]) -> str:
    """Read the Currency column from the Open Positions summary block (D5).

    The summary block header is ``Product | Metric | Amount | Currency``.
    Returns the first non-empty Currency value among the block's data rows,
    or ``""`` if the block is absent or has no Currency value (guard 4 raises
    on this empty result at the call site).
    """
    for i, row in enumerate(rows):
        if (
            len(row) >= 4
            and normalize_header(row[0]) == "product"
            and normalize_header(row[1]) == "metric"
            and normalize_header(row[2]) == "amount"
            and normalize_header(row[3]) == "currency"
        ):
            # Data rows follow the header until the Note/empty row.
            for data_row in rows[i + 1 :]:
                if not data_row or data_row[0] is None:
                    break
                if normalize_header(_str_cell(data_row[0])) == "note":
                    break
                if len(data_row) >= 4:
                    ccy = _str_cell(data_row[3])
                    if ccy:
                        return ccy
            return ""
    return ""


def _find_detail_header(
    rows: list[list[Any]],
    col1_options: tuple[str, ...],
) -> int | None:
    """Find the detail-header row whose col 0 is "Product" and col 1 matches.

    ``col1_options`` is a tuple of normalized candidate values for column 1
    (the Open Positions detail header uses "Instrument/Position"; the
    Closed Positions header uses "Ticker").
    """
    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        if normalize_header(row[0]) != "product":
            continue
        if len(row) < 2:
            continue
        if normalize_header(row[1]) in col1_options:
            return i
    return None


# ---------------------------------------------------------------------------
# Cash Operations (D2/D7/D10/D11/D22, guard 1).
# ---------------------------------------------------------------------------

# Header signature: row[0]=="Type" and row[1]=="Instrument".
_CO_HEADER_COL1 = ("instrument",)


def _parse_cash_operations(
    rows: list[list[Any]],
    account_id: str,
) -> tuple[list[XtbCashOperation], float | None]:
    """Parse Cash Operations -> (operations, free_cash).

    Extracts only ``Type, Ticker, Time, Amount, ID, Comment, Position ID``
    (Instrument/Category/Product dropped). The ``Total`` row's Amount is read
    into ``free_cash`` (2dp) BEFORE excluding it (D22). Total rows are
    excluded from ``cash_operations`` (D10). ``Subaccount transfer`` rows are
    filtered out entirely here (D7 — single place). ``position_id`` is string-
    coerced (guard 1). Amounts rounded to 2dp (D11).
    """
    header_index = _find_cash_header(rows)
    if header_index is None:
        return [], None

    operations: list[XtbCashOperation] = []
    free_cash: float | None = None
    for row in rows[header_index + 1 :]:
        if not row or row[0] is None:
            continue
        op_type = _str_cell(row[0])
        if not op_type:
            continue
        # D10: Total row -> read free_cash then exclude from events.
        if normalize_header(op_type) == "total":
            if len(row) > 5:
                free_cash = round(as_float(row[5]), 2)
            continue
        # D7: filter out Subaccount transfer rows entirely (internal moves).
        if normalize_header(op_type) == "subaccount transfer":
            continue
        time = _to_utc_datetime(row[4] if len(row) > 4 else None)
        if time is None:
            continue
        amount = round(as_float(row[5]) if len(row) > 5 else 0.0, 2)
        operations.append(
            XtbCashOperation(
                account_id=account_id,
                operation_type=op_type,
                ticker=_str_cell(row[2]) if len(row) > 2 else "",
                time=time,
                amount=amount,
                operation_id=_str_cell(row[6]) if len(row) > 6 else "",
                comment=_str_cell(row[7]) if len(row) > 7 else "",
                position_id=_str_cell(row[9]) if len(row) > 9 else "",
            )
        )
    return operations, free_cash


def _find_cash_header(rows: list[list[Any]]) -> int | None:
    """Find the Cash Operations header row (Type | Instrument | Ticker | ...)."""
    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        if (
            normalize_header(row[0]) == "type"
            and normalize_header(row[1]) in _CO_HEADER_COL1
        ):
            return i
    return None


# ---------------------------------------------------------------------------
# Closed Positions (D2/D8/D11, guard 1).
# ---------------------------------------------------------------------------


def _find_closed_header(rows: list[list[Any]]) -> int | None:
    """Find the Closed Positions header row (Instrument | Ticker | Category | ...)."""
    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        if (
            normalize_header(row[0]) == "instrument"
            and normalize_header(row[1]) == "ticker"
        ):
            return i
    return None


def _parse_closed_positions(rows: list[list[Any]]) -> list[XtbClosedPosition]:
    """Parse Closed Positions -> fee-enrichment rows.

    Extracts only ``Position ID, Commission, Purchase value, Sale value,
    Close time`` (the fee-enrichment fields, D8). The ``Profit/loss`` total
    row is excluded. ``position_id`` is string-coerced (guard 1). Commission /
    Purchase value / Sale value rounded to 2dp (D11).
    """
    header_index = _find_closed_header(rows)
    if header_index is None:
        return []

    positions: list[XtbClosedPosition] = []
    for row in rows[header_index + 1 :]:
        if not row or row[0] is None:
            continue
        # Column layout (0-based):
        # 0 Instrument | 1 Ticker | ... | 8 Close Time (UTC) | ... |
        # 12 Purchase Value | 13 Sale Value | 16 Commission | 23 Position ID
        first = normalize_header(_str_cell(row[0]))
        if first == "profit/loss" or first == "total":
            continue
        position_id = _str_cell(row[23]) if len(row) > 23 else ""
        if not position_id:
            continue
        close_time = _to_utc_datetime(row[8] if len(row) > 8 else None)
        if close_time is None:
            continue
        positions.append(
            XtbClosedPosition(
                position_id=position_id,
                commission=round(as_float(row[16]) if len(row) > 16 else 0.0, 2),
                purchase_value=round(as_float(row[12]) if len(row) > 12 else 0.0, 2),
                sale_value=round(as_float(row[13]) if len(row) > 13 else 0.0, 2),
                close_time=close_time,
            )
        )
    return positions


# ---------------------------------------------------------------------------
# Top-level entry point.
# ---------------------------------------------------------------------------


def parse_report(data: bytes, account_id_override: str | None = None) -> XtbReport:
    """Parse a new-format XTB .xlsx report (all 3 sheets) into an ``XtbReport``.

    Account ID comes from R1 ``Account number`` of each sheet; an explicit
    ``account_id_override`` takes precedence. Missing sheets yield an empty
    list for that sheet (guard 5) rather than aborting the whole parse. The
    account currency comes from the Open Positions summary block (D5) and
    raises ``XtbError`` if empty/absent (guard 4).
    """
    wb = _load_workbook(data)

    open_rows = _sheet_rows(wb, "OPEN POSITION")
    cash_rows = _sheet_rows(wb, "CASH OPERATION")
    closed_rows = _sheet_rows(wb, "CLOSED POSITION")

    # Account ID: prefer override, then Open Positions R1, Cash, Closed.
    if account_id_override:
        account_id = account_id_override
    else:
        account_id = (
            _account_id_from_rows(open_rows)
            or _account_id_from_rows(cash_rows)
            or _account_id_from_rows(closed_rows)
            or ""
        )

    # Open Positions (also supplies account_ccy; guard 4 raises if missing
    # AND the sheet is present). Guard 5: a missing sheet -> empty list.
    if open_rows is not None:
        open_positions, account_ccy = _parse_open_positions(open_rows, account_id)
    else:
        open_positions, account_ccy = [], ""

    if cash_rows is not None:
        cash_operations, free_cash = _parse_cash_operations(cash_rows, account_id)
    else:
        cash_operations, free_cash = [], None

    if closed_rows is not None:
        closed_positions = _parse_closed_positions(closed_rows)
    else:
        closed_positions = []

    return XtbReport(
        account_id=account_id,
        account_ccy=account_ccy,
        open_positions=open_positions,
        closed_positions=closed_positions,
        cash_operations=cash_operations,
        free_cash=free_cash,
    )
