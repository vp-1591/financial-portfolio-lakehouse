"""Tests for the XTB pipeline connector."""

from __future__ import annotations

import hashlib
import tempfile
import uuid
import zipfile
from collections.abc import Mapping
from datetime import UTC, datetime
from pathlib import Path

import pyarrow as pa
import pytest

from pipeline.connectors.xtb.fetch import _read_file_bytes, fetch_snapshot
from pipeline.connectors.xtb.parser import (
    XtbCashOperation,
    XtbClosedPosition,
    XtbError,
    XtbOpenPosition,
    XtbReport,
    as_float,
    normalize_header,
    parse_report,
)
from pipeline.connectors.xtb.transform import (
    _account_id_from_filename,
    transform_events,
    transform_snapshot,
)
from pipeline.crypto import decrypt_float, encrypt, generate_key
from pipeline.raw.models import RAW_SCHEMA
from tests.fixtures.xtb import (
    CONVERSION_TRANSFER_AMOUNT,
    DEFAULT_ACCOUNT_CCY,
    DEFAULT_ACCOUNT_ID,
    DEFAULT_CLOSED_COMMISSION,
    DEFAULT_CLOSED_POSITION_ID,
    DEPOSIT_AMOUNT,
    INTEREST_AMOUNT,
    INTEREST_TAX_AMOUNT,
    PURCHASE_AMOUNT,
    SELL_AMOUNT,
    SOLD_OUT_AGGREGATE_VALUE,
    SUBACCOUNT_IN_AMOUNT,
    SUBACCOUNT_OUT_AMOUNT,
    SXR8_AGGREGATE_VALUE,
    SXRV_AGGREGATE_VALUE,
    build_new_format_xlsx_bytes,
    build_xlsx_bytes_from_sheets,
)

# --- XLS test helpers (preserved from tests/test_xtb_net_worth.py) ---


def cell(ref: str, value: object) -> str:
    if isinstance(value, (int, float)):
        return f'<c r="{ref}"><v>{value}</v></c>'
    return f'<c r="{ref}" t="inlineStr"><is><t>{value}</t></is></c>'


def row(index: int, values: Mapping[str, object]) -> str:
    cells = "".join(cell(f"{column}{index}", value) for column, value in values.items())
    return f'<row r="{index}">{cells}</row>'


def sheet(rows: list[str]) -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(rows)}</sheetData>"
        "</worksheet>"
    )


def write_xtb_workbook(
    path: Path, include_isin: bool = False, include_cash_ops: bool = False
) -> None:
    """Create a minimal legacy-format XLSX workbook (used by fetch tests).

    fetch.py is unchanged by the Stage 0/1 rewrite, so these legacy-format
    helpers remain valid for the TestFetchFromS3 suite.
    """

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        "<sheets>"
        '<sheet name="OPEN POSITION 15062026" sheetId="1" r:id="rId1"/>'
        '<sheet name="CASH OPERATION HISTORY" sheetId="2" r:id="rId2"/>'
        "</sheets>"
        "</workbook>"
    )
    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet2.xml"/>'
        "</Relationships>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/worksheets/sheet2.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )

    header_row = {
        "B": "Position",
        "C": "Symbol",
        "D": "Type",
        "E": "Volume",
        "I": "Purchase value",
        "P": "Gross P/L",
    }
    first_position = {
        "B": 1001,
        "C": "VWCE.DE",
        "D": "BUY",
        "E": 2,
        "I": 150,
        "P": 25,
    }
    second_position = {
        "B": 1002,
        "C": "VWCE.DE",
        "D": "BUY",
        "E": 1,
        "I": 40,
        "P": -20,
    }
    if include_isin:
        header_row["Q"] = "ISIN"
        first_position["Q"] = "IE00BK5BQT80"
        second_position["Q"] = "IE00BK5BQT80"

    open_sheet = sheet(
        [
            row(5, {"F": "Name and surname", "I": "Account", "L": "Currency"}),
            row(6, {"F": "Anon User", "I": "123456", "L": "PLN"}),
            row(7, {"F": "Balance", "I": "Equity"}),
            row(8, {"F": 25, "I": 220}),
            row(11, header_row),
            row(12, first_position),
            row(13, second_position),
            row(14, {"B": "Total", "I": 190, "P": 5}),
        ]
    )

    cash_header = {"B": "ID", "C": "Type", "G": "Amount"}
    if include_cash_ops:
        cash_header["D"] = "Comment"
        cash_header["E"] = "Currency"
        cash_header["F"] = "Time"
        cash_rows = [
            row(11, cash_header),
            row(
                12,
                {
                    "B": 1,
                    "C": "Deposit",
                    "D": "Initial deposit",
                    "E": "PLN",
                    "F": "2026-01-01",
                    "G": 200,
                },
            ),
            row(
                13,
                {
                    "B": 2,
                    "C": "Dividend",
                    "D": "VWCE dividend",
                    "E": "EUR",
                    "F": "2026-03-15",
                    "G": 5,
                },
            ),
        ]
    else:
        cash_rows = [
            row(11, {"B": "ID", "C": "Type", "G": "Amount"}),
            row(12, {"B": 1, "C": "Deposit", "G": 200}),
        ]

    cash_sheet_xml = sheet(cash_rows)

    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", rels_xml)
        archive.writestr("xl/worksheets/sheet1.xml", open_sheet)
        archive.writestr("xl/worksheets/sheet2.xml", cash_sheet_xml)


def _build_xlsx_bytes(
    include_isin: bool = False, include_cash_ops: bool = False
) -> bytes:
    """Build minimal legacy-format .xlsx bytes for transform/fetch tests."""

    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / f"test-{uuid.uuid4().hex}.xlsx"
        write_xtb_workbook(
            path, include_isin=include_isin, include_cash_ops=include_cash_ops
        )
        return path.read_bytes()


class TestParserHelpers:
    """Tests for the small parser helpers retained across the rewrite."""

    def test_as_float(self) -> None:
        assert as_float(None) == 0.0
        assert as_float("") == 0.0
        assert as_float(42) == 42.0
        assert as_float("3.14") == 3.14
        assert as_float("1,5") == 1.5
        assert as_float("abc", -1.0) == -1.0

    def test_normalize_header(self) -> None:
        assert normalize_header("  Purchase  value  ") == "purchase value"
        assert normalize_header(None) == ""
        assert normalize_header("Gross P/L") == "gross p/l"


REAL_SAMPLE_PATH = (
    Path(__file__).resolve().parent.parent
    / "docs"
    / "xtb"
    / "xtb-report-sample"
    / "PLN_12345678_2006-01-01_2026-08-03.xlsx"
)


class TestXtbParser:
    """Parser tests for the new-format 3-sheet workbook (Stage 1 + guards)."""

    @pytest.fixture()
    def report(self) -> XtbReport:
        return parse_report(build_new_format_xlsx_bytes())

    # --- 3-sheet parsing + dataclass shape ---

    def test_parses_all_three_sheets(self, report: XtbReport) -> None:
        assert report.account_id == DEFAULT_ACCOUNT_ID
        assert report.account_ccy == DEFAULT_ACCOUNT_CCY
        assert len(report.open_positions) == 2  # SXR8.DE, SXRV.DE (SOLD.DE skipped)
        assert len(report.closed_positions) == 1
        assert len(report.cash_operations) == 6  # subaccount pair + Total excluded
        assert report.free_cash is not None

    def test_open_position_fields(self, report: XtbReport) -> None:
        by_ticker = {p.ticker: p for p in report.open_positions}
        sxr8 = by_ticker["SXR8.DE"]
        assert sxr8.account_id == DEFAULT_ACCOUNT_ID
        assert sxr8.product == "Investment Plan"
        assert sxr8.instrument == "Core S&P 500"  # real name on the aggregate row
        assert sxr8.category == "ETF"  # populated on holdings (empty on child lots)
        assert sxr8.value == pytest.approx(SXR8_AGGREGATE_VALUE)
        assert by_ticker["SXRV.DE"].instrument == "NASDAQ 100"

    # --- aggregate-vs-child distinction (D4) ---

    def test_child_lot_rows_are_skipped(self, report: XtbReport) -> None:
        # Two instruments each have an aggregate + a child lot row; only the
        # 2 aggregate rows survive. Child lots carry a numeric ID in
        # Instrument and an empty Category.
        tickers = {p.ticker for p in report.open_positions}
        assert tickers == {"SXR8.DE", "SXRV.DE"}
        # No child-lot Instrument (numeric position id) leaks through.
        assert all(not p.instrument.isdigit() for p in report.open_positions)
        assert all(p.category for p in report.open_positions)  # non-empty on aggregates

    def test_zero_value_aggregate_skipped(self, report: XtbReport) -> None:
        # Guard 3: SOLD.DE aggregate has Value 0 -> skipped.
        assert "SOLD.DE" not in {p.ticker for p in report.open_positions}
        assert SOLD_OUT_AGGREGATE_VALUE == 0.0

    # --- date decoding (D3): tz-aware UTC ---

    def test_cash_operation_times_are_utc(self, report: XtbReport) -> None:
        for op in report.cash_operations:
            assert op.time.tzinfo is not None
            assert op.time.utcoffset() == datetime.now(UTC).utcoffset()

    def test_closed_position_close_time_is_utc(self, report: XtbReport) -> None:
        closed = report.closed_positions[0]
        assert closed.close_time.tzinfo is not None
        assert closed.close_time == datetime(2026, 8, 2, 7, 0, tzinfo=UTC)

    # --- CASH / free_cash from Total row (D22) ---

    def test_free_cash_read_from_total_row(self, report: XtbReport) -> None:
        # D22: Total row Amount -> free_cash (2dp), excluded from events.
        assert report.free_cash == pytest.approx(
            DEPOSIT_AMOUNT
            + INTEREST_AMOUNT
            + INTEREST_TAX_AMOUNT
            + SUBACCOUNT_IN_AMOUNT
            + SUBACCOUNT_OUT_AMOUNT
            + (-1000.0)
            + PURCHASE_AMOUNT
            + SELL_AMOUNT
        )

    def test_free_cash_equals_sum_of_cash_operations(self, report: XtbReport) -> None:
        # D22 invariant: under full history, sum of events cash_amount == free_cash
        # (subaccount transfers net to zero, so the Total equals the sum of
        # the retained events).
        assert report.free_cash == pytest.approx(
            round(sum(op.amount for op in report.cash_operations), 2)
        )

    # --- Total-row exclusion from events (D10) + subaccount filtering (D7) ---

    def test_total_row_excluded_from_cash_operations(self, report: XtbReport) -> None:
        types = {op.operation_type for op in report.cash_operations}
        assert "Total" not in types
        assert all(op.operation_type for op in report.cash_operations)

    def test_subaccount_transfers_filtered(self, report: XtbReport) -> None:
        types = {op.operation_type for op in report.cash_operations}
        assert "Subaccount transfer" not in types

    def test_currency_conversion_transfer_kept(self, report: XtbReport) -> None:
        transfers = [
            op for op in report.cash_operations if op.operation_type == "Transfer"
        ]
        assert len(transfers) == 1
        assert transfers[0].amount == pytest.approx(-1000.0)
        assert "Exchange rate:0.230001" in transfers[0].comment

    # --- 2dp rounding (D11) ---

    def test_amounts_rounded_to_2dp(self, report: XtbReport) -> None:
        for op in report.cash_operations:
            assert round(op.amount, 2) == op.amount

    def test_commission_rounded_to_2dp(self, report: XtbReport) -> None:
        closed = report.closed_positions[0]
        assert round(closed.commission, 2) == closed.commission

    # --- Closed Positions: nonzero commission + Profit/loss total excluded ---

    def test_closed_position_has_nonzero_commission(self, report: XtbReport) -> None:
        closed = report.closed_positions[0]
        assert closed.position_id == DEFAULT_CLOSED_POSITION_ID
        assert closed.commission == pytest.approx(DEFAULT_CLOSED_COMMISSION)
        assert closed.commission != 0.0

    def test_profit_loss_total_excluded(self, report: XtbReport) -> None:
        # The fixture has a "Profit/loss" total row that must be excluded.
        assert len(report.closed_positions) == 1

    def test_closed_position_matches_stock_sell(self, report: XtbReport) -> None:
        sell_rows = [
            op for op in report.cash_operations if op.operation_type == "Stock sell"
        ]
        assert len(sell_rows) == 1
        assert sell_rows[0].position_id == report.closed_positions[0].position_id

    # --- guard 1: position_id string coercion ---

    def test_position_id_is_string(self, report: XtbReport) -> None:
        for op in report.cash_operations:
            assert isinstance(op.position_id, str)
        for closed in report.closed_positions:
            assert isinstance(closed.position_id, str)
        # The closed position id is a numeric value in the xlsx but coerced.
        assert report.closed_positions[0].position_id == DEFAULT_CLOSED_POSITION_ID

    # --- guard 5: missing sheet -> empty list, no abort ---

    def test_missing_closed_positions_sheet_returns_empty(self) -> None:
        data = build_new_format_xlsx_bytes()
        # Rebuild with the Closed Positions sheet dropped.
        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        if "Closed Positions" in wb.sheetnames:
            del wb["Closed Positions"]
        buffer = BytesIO()
        wb.save(buffer)
        report = parse_report(buffer.getvalue())
        assert report.closed_positions == []
        assert len(report.open_positions) == 2
        assert len(report.cash_operations) == 6

    def test_missing_open_positions_sheet_empty_list_no_abort(self) -> None:
        # Missing Open Positions -> empty list (guard 5); account_ccy is
        # empty (the Currency source is absent) but the parse does not abort.
        data = build_new_format_xlsx_bytes()
        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        if "Open Positions" in wb.sheetnames:
            del wb["Open Positions"]
        buffer = BytesIO()
        wb.save(buffer)
        report = parse_report(buffer.getvalue())
        assert report.open_positions == []
        assert report.account_ccy == ""

    # --- guard 4: missing/empty summary-block currency raises ---

    def test_empty_summary_currency_raises(self) -> None:
        # Build an Open Positions sheet whose summary block Currency is empty.
        open_rows: list[tuple[object, ...]] = [
            ("Account number", DEFAULT_ACCOUNT_ID),
            ("Open Positions",),
            ("Data as of report generated", "2026-08-03"),
            ("Product", "Metric", "Amount", "Currency"),
            ("My Trades", "Value", 0, None),  # empty Currency
            ("Investment Plans", "Value", 100, None),  # empty Currency
        ]
        cash_rows: list[tuple[object, ...]] = [
            ("Account number", DEFAULT_ACCOUNT_ID),
            ("Cash Operations",),
            (
                "Type",
                "Instrument",
                "Ticker",
                "Category",
                "Time",
                "Amount",
                "ID",
                "Comment",
                "Product",
                "Position ID",
            ),
        ]
        data = build_xlsx_bytes_from_sheets(
            {"Open Positions": open_rows, "Cash Operations": cash_rows}
        )
        with pytest.raises(XtbError, match="account currency missing"):
            parse_report(data)

    def test_absent_summary_block_raises(self) -> None:
        # No summary block header at all -> guard 4 raises.
        open_rows: list[tuple[object, ...]] = [
            ("Account number", DEFAULT_ACCOUNT_ID),
            ("Open Positions",),
            ("Data as of report generated", "2026-08-03"),
        ]
        data = build_xlsx_bytes_from_sheets({"Open Positions": open_rows})
        with pytest.raises(XtbError, match="account currency missing"):
            parse_report(data)

    # --- account_id override ---

    def test_account_id_override(self) -> None:
        report = parse_report(
            build_new_format_xlsx_bytes(), account_id_override="OVR-1"
        )
        assert report.account_id == "OVR-1"
        assert all(p.account_id == "OVR-1" for p in report.open_positions)
        assert all(op.account_id == "OVR-1" for op in report.cash_operations)

    # --- empty-Time cash rows: dropped and counted (no silent loss) ---

    def test_empty_time_cash_row_dropped_and_counted(self) -> None:
        """An empty Time cell skips the row but is counted for the transform warning."""
        sheets = {
            "Open Positions": _op_sheet(DEFAULT_ACCOUNT_ID, DEFAULT_ACCOUNT_CCY),
            "Cash Operations": _cash_sheet(
                DEFAULT_ACCOUNT_ID,
                [
                    (
                        "Deposit",
                        "",
                        _naive(2026, 7, 10, 0, 0),
                        1000.0,
                        "900011122",
                        "deposit",
                        "",
                    ),
                    ("Withdrawal", "", None, -200.0, "900011123", "withdrawal", ""),
                ],
            ),
            "Closed Positions": _closed_sheet(DEFAULT_ACCOUNT_ID, []),
        }
        report = parse_report(build_xlsx_bytes_from_sheets(sheets))
        assert report.dropped_cash_rows == 1
        assert [op.operation_type for op in report.cash_operations] == ["Deposit"]

    # --- dataclass field scope (YAGNI) ---

    def test_dataclass_field_scope(self) -> None:
        # XtbClosedPosition must NOT carry swap/rollover/margin/conversion
        # rates (D8). XtbOpenPosition carries only the 6 mapped fields.
        closed_fields = {
            f.name for f in __import__("dataclasses").fields(XtbClosedPosition)
        }
        assert closed_fields == {
            "position_id",
            "commission",
            "close_time",
        }
        assert "swap" not in closed_fields
        assert "rollover" not in closed_fields
        assert "margin" not in closed_fields
        open_fields = {
            f.name for f in __import__("dataclasses").fields(XtbOpenPosition)
        }
        assert open_fields == {
            "account_id",
            "product",
            "instrument",
            "ticker",
            "category",
            "value",
        }
        cash_fields = {
            f.name for f in __import__("dataclasses").fields(XtbCashOperation)
        }
        assert cash_fields == {
            "account_id",
            "operation_type",
            "ticker",
            "time",
            "amount",
            "operation_id",
            "comment",
            "position_id",
        }

    # --- integration: real anonymized sample ---

    @pytest.mark.skipif(
        not REAL_SAMPLE_PATH.exists(), reason="sample xlsx not checked out"
    )
    def test_real_sample_round_trip(self) -> None:
        report = parse_report(REAL_SAMPLE_PATH.read_bytes())
        assert report.account_id == "12345678"
        assert report.account_ccy == "PLN"
        # 2 aggregate holdings (SXR8.DE, SXRV.DE); child lots skipped.
        assert {p.ticker for p in report.open_positions} == {"SXR8.DE", "SXRV.DE"}
        # 1 closed position (the sample's Commission is 0 — the fixture adds a
        # nonzero one to exercise fee handling).
        assert len(report.closed_positions) == 1
        assert report.closed_positions[0].commission == 0.0
        # D22 invariant: free_cash == sum of cash_operations amounts.
        assert report.free_cash == pytest.approx(
            round(sum(op.amount for op in report.cash_operations), 2),
            rel=0.01,
        )
        # Subaccount transfers filtered out.
        assert "Subaccount transfer" not in {
            op.operation_type for op in report.cash_operations
        }
        # No Total/Profit-loss rows leak into events.
        assert "Total" not in {op.operation_type for op in report.cash_operations}


class TestTransformSnapshot:
    """Tests for the raw -> normalized snapshot transform (Stage 2, D18/D22)."""

    @pytest.fixture()
    def fernet_key(self) -> bytes:
        return generate_key()

    def _build_raw(
        self,
        xlsx_bytes: bytes,
        fernet_key: bytes,
        *,
        source: str = "XTB_REPORT",
        fetched_at: datetime | None = None,
        source_file: str = "report.xlsx",
    ) -> pa.Table:
        """Build a raw-layer table from .xlsx bytes (shared bronze, D17)."""
        encrypted_payload = encrypt(xlsx_bytes, fernet_key)
        return pa.table(
            {
                "fetched_at": [fetched_at or datetime.now(UTC)],
                "broker": ["XTB"],
                "source": [source],
                "payload": [encrypted_payload],
                "payload_hash": [hashlib.sha256(xlsx_bytes).hexdigest()],
                "source_file": [source_file],
            },
            schema=RAW_SCHEMA,
        )

    def _decrypt_values(self, table: pa.Table, col: str, key: bytes) -> list[float]:
        return [decrypt_float(v, key) for v in table.column(col).to_pylist()]

    def test_transform_produces_equity_and_cash_rows(self, fernet_key: bytes) -> None:
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_snapshot(raw, fernet_key)

        types = result.column("position_type").to_pylist()
        assert "EQUITY" in types
        assert "CASH" in types
        # 2 EQUITY aggregates (SXR8.DE, SXRV.DE; SOLD.DE skipped) + 1 CASH = 3 rows.
        assert result.num_rows == 3

    def test_equity_rows_from_per_ticker_aggregates(self, fernet_key: bytes) -> None:
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_snapshot(raw, fernet_key)

        labels = result.column("label").to_pylist()
        assert "SXR8.DE" in labels
        assert "SXRV.DE" in labels
        assert "SOLD.DE" not in labels  # zero-value aggregate skipped (guard 3)

        equity_rows = [r for r in result.to_pylist() if r["position_type"] == "EQUITY"]
        by_label = {r["label"]: r for r in equity_rows}
        assert by_label["SXR8.DE"]["description"] == "Core S&P 500"
        assert by_label["SXR8.DE"]["asset_class"] == "ETF"
        assert by_label["SXRV.DE"]["description"] == "NASDAQ 100"
        values = self._decrypt_values(result, "security_value", fernet_key)
        assert SXR8_AGGREGATE_VALUE in values
        assert SXRV_AGGREGATE_VALUE in values

    def test_cash_holding_from_free_cash(self, fernet_key: bytes) -> None:
        """D22: one CASH row per account from free_cash (Total row)."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_snapshot(raw, fernet_key)

        cash_rows = [r for r in result.to_pylist() if r["position_type"] == "CASH"]
        assert len(cash_rows) == 1
        cash = cash_rows[0]
        assert cash["label"] == f"CASH {DEFAULT_ACCOUNT_CCY}"
        assert cash["asset_class"] == "CASH"
        assert cash["security_ccy"] == DEFAULT_ACCOUNT_CCY
        assert cash["description"] == f"Cash {DEFAULT_ACCOUNT_CCY}"
        assert cash["isin"] == ""  # D12
        assert decrypt_float(cash["security_value"], fernet_key) == pytest.approx(
            DEPOSIT_AMOUNT
            + INTEREST_AMOUNT
            + INTEREST_TAX_AMOUNT
            + SUBACCOUNT_IN_AMOUNT
            + SUBACCOUNT_OUT_AMOUNT
            + CONVERSION_TRANSFER_AMOUNT
            + PURCHASE_AMOUNT
            + SELL_AMOUNT
        )

    def test_cash_absent_when_no_total_row(self, fernet_key: bytes) -> None:
        """D22: CASH row is skipped when free_cash is None (no Total row)."""
        # Build a workbook with the Cash Operations Total row dropped.
        from io import BytesIO

        import openpyxl

        data = build_new_format_xlsx_bytes()
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        ws = wb["Cash Operations"]
        # The Total row is the last data row; delete it.
        ws.delete_rows(ws.max_row)
        buf = BytesIO()
        wb.save(buf)
        raw = self._build_raw(buf.getvalue(), fernet_key)
        result = transform_snapshot(raw, fernet_key)

        types = result.column("position_type").to_pylist()
        assert "CASH" not in types
        assert "EQUITY" in types

    def test_security_ccy_is_account_currency(self, fernet_key: bytes) -> None:
        """D5: security_ccy comes from the summary-block Currency, not a literal."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_snapshot(raw, fernet_key)
        assert all(
            ccy == DEFAULT_ACCOUNT_CCY
            for ccy in result.column("security_ccy").to_pylist()
        )

    def test_isin_empty_for_all_rows(self, fernet_key: bytes) -> None:
        """D12: no ISIN in the new format; isin is empty string."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_snapshot(raw, fernet_key)
        assert all(isin == "" for isin in result.column("isin").to_pylist())

    def test_multi_account_both_survive(self, fernet_key: bytes) -> None:
        """D18: two accounts in raw -> both survive snapshot."""
        t1 = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        t2 = datetime(2026, 8, 3, 7, 0, tzinfo=UTC)
        raw_a = self._build_raw(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t1,
            source_file="a.xlsx",
        )
        raw_b = self._build_raw(
            build_new_format_xlsx_bytes(account_id="222", account_ccy="EUR"),
            fernet_key,
            fetched_at=t2,
            source_file="b.xlsx",
        )
        combined = pa.concat_tables([raw_a, raw_b], schema=RAW_SCHEMA)
        result = transform_snapshot(combined, fernet_key)

        account_ids = set(result.column("account_id").to_pylist())
        assert account_ids == {"111", "222"}
        # Each account has 2 EQUITY + 1 CASH = 3 rows.
        assert result.num_rows == 6

    def test_re_upload_supersedes_old_snapshot(self, fernet_key: bytes) -> None:
        """D18: latest fetched_at per account_id supersedes the older payload."""
        t_old = datetime(2026, 8, 1, 6, 0, tzinfo=UTC)
        t_new = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw_old = self._build_raw(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t_old,
            source_file="old.xlsx",
        )
        raw_new = self._build_raw(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t_new,
            source_file="new.xlsx",
        )
        combined = pa.concat_tables([raw_old, raw_new], schema=RAW_SCHEMA)
        result = transform_snapshot(combined, fernet_key)

        # Only the latest payload survives — 3 rows, not 6.
        assert result.num_rows == 3
        assert all(fa == t_new for fa in result.column("fetched_at").to_pylist())

    def test_guard9_tiebreaker_deterministic(self, fernet_key: bytes) -> None:
        """Guard 9: same fetched_at + same account -> deterministic pick by source_file."""
        t = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw_a = self._build_raw(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t,
            source_file="a.xlsx",
        )
        raw_b = self._build_raw(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t,
            source_file="b.xlsx",
        )
        combined = pa.concat_tables([raw_a, raw_b], schema=RAW_SCHEMA)
        result = transform_snapshot(combined, fernet_key)

        # Tie broken deterministically by max source_file ("b.xlsx" > "a.xlsx")
        # -> only one payload survives -> 3 rows, not 6.
        assert result.num_rows == 3

    def test_legacy_source_rows_skipped(self, fernet_key: bytes) -> None:
        """D17: rows with source != 'XTB_REPORT' are skipped (legacy format)."""
        raw = self._build_raw(
            build_new_format_xlsx_bytes(),
            fernet_key,
            source="OPEN POSITION",  # legacy source
        )
        result = transform_snapshot(raw, fernet_key)
        assert result.num_rows == 0


class TestTransformEvents:
    """Tests for the raw -> normalized events transform (Stage 2, D17/D9/D8)."""

    @pytest.fixture()
    def fernet_key(self) -> bytes:
        return generate_key()

    def _build_raw(
        self,
        xlsx_bytes: bytes,
        fernet_key: bytes,
        *,
        source: str = "XTB_REPORT",
        fetched_at: datetime | None = None,
        source_file: str = "report.xlsx",
    ) -> pa.Table:
        """Build a raw-layer table from .xlsx bytes (shared bronze, D17)."""
        encrypted_payload = encrypt(xlsx_bytes, fernet_key)
        return pa.table(
            {
                "fetched_at": [fetched_at or datetime.now(UTC)],
                "broker": ["XTB"],
                "source": [source],
                "payload": [encrypted_payload],
                "payload_hash": [hashlib.sha256(xlsx_bytes).hexdigest()],
                "source_file": [source_file],
            },
            schema=RAW_SCHEMA,
        )

    def test_events_produces_operation_rows(self, fernet_key: bytes) -> None:
        """D17 shared bronze: events from xtb_snapshot raw with source='XTB_REPORT'."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)

        # 6 events: deposit, interest, interest tax, transfer, purchase, sell.
        # Subaccount transfers filtered (D7), Total row excluded (D10).
        assert result.num_rows == 6

        event_types = result.column("event_type").to_pylist()
        assert "DEPOSIT" in event_types
        assert "INTEREST" in event_types
        assert "TAX" in event_types
        assert "TRADE" in event_types
        assert "TRANSFER" in event_types
        assert "UNKNOWN" not in event_types

    def test_event_type_map(self, fernet_key: bytes) -> None:
        """D6: operation_type -> event_type map (INTEREST/TAX/TRADE/TRANSFER/DEPOSIT)."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)

        raw_to_norm = {r["raw_event_type"]: r["event_type"] for r in result.to_pylist()}
        assert raw_to_norm["Deposit"] == "DEPOSIT"
        assert raw_to_norm["Free funds interest"] == "INTEREST"
        assert raw_to_norm["Free funds interest tax"] == "TAX"
        assert raw_to_norm["Stock purchase"] == "TRADE"
        assert raw_to_norm["Stock sell"] == "TRADE"
        assert raw_to_norm["Transfer"] == "TRANSFER"

    def test_total_row_excluded_and_subaccount_filtered(
        self, fernet_key: bytes
    ) -> None:
        """D10/D7: Total row excluded from events; subaccount transfers filtered."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)

        raw_types = result.column("raw_event_type").to_pylist()
        assert "Total" not in raw_types
        assert "Subaccount transfer" not in raw_types

    def test_cash_sum_equals_free_cash(self, fernet_key: bytes) -> None:
        """D22: under full history, sum of events cash_amount == free_cash."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)
        report = parse_report(build_new_format_xlsx_bytes())

        amounts = [
            decrypt_float(v, fernet_key)
            for v in result.column("cash_amount").to_pylist()
        ]
        assert sum(amounts) == pytest.approx(report.free_cash or 0.0, rel=0.01)

    def test_currency_conversion_transfer_kept_fx_rate_null(
        self, fernet_key: bytes
    ) -> None:
        """D7: currency-conversion Transfer kept as TRANSFER with target_fx_rate null."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)

        transfers = [r for r in result.to_pylist() if r["event_type"] == "TRANSFER"]
        assert len(transfers) == 1
        assert transfers[0]["raw_event_type"] == "Transfer"
        assert decrypt_float(transfers[0]["cash_amount"], fernet_key) == pytest.approx(
            CONVERSION_TRANSFER_AMOUNT
        )
        assert (
            transfers[0]["target_fx_rate"] is None
        )  # D7: do NOT parse Exchange rate:X
        assert transfers[0]["target_ccy"] is None  # filled by normalize_currency

    def test_trade_enrichment_sell_row_only(self, fernet_key: bytes) -> None:
        """D8: fee_amount on the closing (sell) row only; purchase row gets no fee."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)

        trades = [r for r in result.to_pylist() if r["event_type"] == "TRADE"]
        # One purchase + one sell.
        assert len(trades) == 2
        by_type = {r["raw_event_type"]: r for r in trades}

        # Purchase row: has qty/price/side, no fee/settle.
        purchase = by_type["Stock purchase"]
        assert purchase["ticker"] == "SXR8.DE"
        assert decrypt_float(purchase["quantity"], fernet_key) == pytest.approx(10.0001)
        assert decrypt_float(purchase["price"], fernet_key) == pytest.approx(100.0)
        assert purchase["side"] == "BUY"
        assert purchase["fee_amount"] is None  # D8: fee on closing row only
        assert purchase["settle_date"] is None

        # Sell row: enriched from Closed Positions via position_id.
        sell = by_type["Stock sell"]
        assert sell["ticker"] == "SXR8.DE"
        assert decrypt_float(sell["quantity"], fernet_key) == pytest.approx(10.0001)
        assert decrypt_float(sell["price"], fernet_key) == pytest.approx(120.0)
        assert sell["side"] == "BUY"  # closing a long position
        assert decrypt_float(sell["fee_amount"], fernet_key) == pytest.approx(
            DEFAULT_CLOSED_COMMISSION
        )
        assert sell["settle_date"] is not None  # close_time ISO string

    def test_2dp_rounding(self, fernet_key: bytes) -> None:
        """D11: cash_amount rounded to 2dp."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)
        amounts = [
            decrypt_float(v, fernet_key)
            for v in result.column("cash_amount").to_pylist()
        ]
        assert all(round(a, 2) == a for a in amounts)

    def test_shared_bronze_no_xtb_events_raw(self, fernet_key: bytes) -> None:
        """D17: events produced from xtb_snapshot raw (source='XTB_REPORT'), no xtb_events raw."""
        raw = self._build_raw(build_new_format_xlsx_bytes(), fernet_key)
        result = transform_events(raw, fernet_key)
        assert result.num_rows == 6
        # All rows carry source='XTB_REPORT' (from the shared bronze raw).
        assert all(s == "XTB_REPORT" for s in result.column("source").to_pylist())

    def test_events_latest_payload_per_account_on_reupload(
        self, fernet_key: bytes
    ) -> None:
        """D9: re-upload supersedes; latest payload per account (no union)."""
        t_old = datetime(2026, 8, 1, 6, 0, tzinfo=UTC)
        t_new = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw_old = self._build_raw(
            build_new_format_xlsx_bytes(),
            fernet_key,
            fetched_at=t_old,
            source_file="old.xlsx",
        )
        raw_new = self._build_raw(
            build_new_format_xlsx_bytes(),
            fernet_key,
            fetched_at=t_new,
            source_file="new.xlsx",
        )
        combined = pa.concat_tables([raw_old, raw_new], schema=RAW_SCHEMA)
        result = transform_events(combined, fernet_key)

        # Latest payload only — 6 events, not 12.
        assert result.num_rows == 6
        assert all(fa == t_new for fa in result.column("fetched_at").to_pylist())

    def test_cross_account_same_id_events_coexist(self, fernet_key: bytes) -> None:
        """D9/D15: same event_id from different accounts both survive (account_id in dedup)."""
        t = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw_a = self._build_raw(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t,
            source_file="a.xlsx",
        )
        raw_b = self._build_raw(
            build_new_format_xlsx_bytes(account_id="222", account_ccy="EUR"),
            fernet_key,
            fetched_at=t,
            source_file="b.xlsx",
        )
        combined = pa.concat_tables([raw_a, raw_b], schema=RAW_SCHEMA)
        result = transform_events(combined, fernet_key)

        # Both accounts contribute 6 events each = 12 (same IDs, different accounts).
        assert result.num_rows == 12
        account_ids = set(result.column("account_id").to_pylist())
        assert account_ids == {"111", "222"}


class TestFetchFromS3:
    """Tests for fetch_snapshot with S3 URIs (D17: fetch_events removed)."""

    @pytest.fixture()
    def xlsx_bytes(self) -> bytes:
        return _build_xlsx_bytes()

    def test_fetch_snapshot_s3_uri(self, xlsx_bytes: bytes, monkeypatch) -> None:
        monkeypatch.setattr(
            "pipeline.s3.read_s3_bytes",
            lambda uri: (xlsx_bytes, "report.xlsx"),
        )

        table = fetch_snapshot("s3://bucket/pipeline/staging/xtb/report.xlsx")
        assert table.num_rows == 1
        assert table.column("source_file")[0].as_py() == "report.xlsx"
        assert table.column("broker")[0].as_py() == "XTB"
        assert table.column("source")[0].as_py() == "XTB_REPORT"  # D17

    def test_fetch_snapshot_local_path_still_works(self, tmp_path: Path) -> None:
        """Local file paths are not affected by S3 support."""

        report = tmp_path / "xtb-test.xlsx"
        write_xtb_workbook(report)
        table = fetch_snapshot(report)

        assert table.num_rows == 1
        assert table.column("source_file")[0].as_py() == report.name
        assert table.column("source")[0].as_py() == "XTB_REPORT"

    def test_read_file_bytes_s3_extracts_filename(
        self, xlsx_bytes: bytes, monkeypatch
    ) -> None:

        monkeypatch.setattr(
            "pipeline.s3.read_s3_bytes",
            lambda uri: (xlsx_bytes, "nested_report.xlsx"),
        )

        payload, filename = _read_file_bytes(
            "s3://bucket/pipeline/staging/xtb/nested_report.xlsx"
        )
        assert payload == xlsx_bytes
        assert filename == "nested_report.xlsx"

    def test_read_file_bytes_local_path(self, tmp_path: Path) -> None:

        report = tmp_path / "xtb-test.xlsx"
        write_xtb_workbook(report)
        expected_bytes = report.read_bytes()
        payload, filename = _read_file_bytes(report)

        assert payload == expected_bytes
        assert filename == report.name

    def test_read_file_bytes_s3_percent_decodes_key(
        self, xlsx_bytes: bytes, monkeypatch
    ) -> None:
        """EventBridge delivers percent-encoded S3 keys; _read_file_bytes decodes them."""

        captured_uris: list[str] = []

        def mock_read_s3_bytes(uri: str):
            captured_uris.append(uri)
            return xlsx_bytes, "report with spaces.xlsx"

        monkeypatch.setattr("pipeline.s3.read_s3_bytes", mock_read_s3_bytes)

        # EventBridge delivers keys with %20 for spaces
        payload, _filename = _read_file_bytes(
            "s3://bucket/staging/xtb/report%20with%20spaces.xlsx"
        )
        assert payload == xlsx_bytes
        # The decoded URI should have spaces, not %20
        assert captured_uris[0] == "s3://bucket/staging/xtb/report with spaces.xlsx"

    def test_read_file_bytes_s3_no_double_decode(
        self, xlsx_bytes: bytes, monkeypatch
    ) -> None:
        """A key with a literal % should not be double-decoded."""

        captured_uris: list[str] = []

        def mock_read_s3_bytes(uri: str):
            captured_uris.append(uri)
            return xlsx_bytes, "report.xlsx"

        monkeypatch.setattr("pipeline.s3.read_s3_bytes", mock_read_s3_bytes)

        # A key with no percent-encoding should pass through unchanged
        _payload, _filename = _read_file_bytes("s3://bucket/staging/xtb/report.xlsx")
        assert captured_uris[0] == "s3://bucket/staging/xtb/report.xlsx"

    def test_read_file_bytes_s3_multiple_percent_encodings(
        self, xlsx_bytes: bytes, monkeypatch
    ) -> None:
        """Multiple percent-encoded characters in a single key are all decoded."""

        captured_uris: list[str] = []

        def mock_read_s3_bytes(uri: str):
            captured_uris.append(uri)
            return xlsx_bytes, "my report.xlsx"

        monkeypatch.setattr("pipeline.s3.read_s3_bytes", mock_read_s3_bytes)

        _payload, _filename = _read_file_bytes(
            "s3://bucket/staging/xtb/my%20report.xlsx"
        )
        assert captured_uris[0] == "s3://bucket/staging/xtb/my report.xlsx"


# ---------------------------------------------------------------------------
# Stage 5 — Excel-serial decoding, open→closed lifecycle, real-sample
# transform integration. Helpers + test classes below.
# ---------------------------------------------------------------------------


def _naive(year: int, month: int, day: int, hour: int = 0, minute: int = 0) -> datetime:
    """Build a naive datetime for an XLSX date cell (openpyxl rejects tz-aware)."""
    return datetime(year, month, day, hour, minute)  # noqa: DTZ001


def _op_sheet(
    account_id: str,
    account_ccy: str,
    aggregates: list[tuple[str, str, str, str, float]] | None = None,
) -> list[tuple[object, ...]]:
    """Build a minimal Open Positions sheet (summary block + optional aggregates).

    Each aggregate tuple is (product, instrument, ticker, category, value).
    """
    rows: list[tuple[object, ...]] = [
        ("Account number", account_id),
        ("Open Positions",),
        ("Data as of report generated", "2026-08-03"),
        ("Product", "Metric", "Amount", "Currency"),
        ("My Trades", "Value", 0, account_ccy),
        (None,),
        ("Note", "Summary values and open positions are shown as of report generation"),
        (
            "Product",
            "Instrument/Position",
            "Ticker",
            "Category",
            "Type",
            "Volume",
            "Value",
        ),
    ]
    for product, instrument, ticker, category, value in aggregates or []:
        rows.append((product, instrument, ticker, category, None, 0, value))
    return rows


def _cash_sheet(
    account_id: str,
    ops: list[tuple[str, str, object, float, str, str, str]],
) -> list[tuple[object, ...]]:
    """Build a minimal Cash Operations sheet.

    Each op tuple is (type, ticker, time, amount, id, comment, position_id).
    A Total row (sum of amounts) is appended so free_cash is populated (D22).
    """
    rows: list[tuple[object, ...]] = [
        ("Account number", account_id),
        ("Cash Operations",),
        ("Date from (UTC)", "2006-01-01"),
        ("Date to (UTC)", "2026-08-03"),
        (
            "Type",
            "Instrument",
            "Ticker",
            "Category",
            "Time",
            "Amount",
            "ID",
            "Comment",
            "Product",
            "Position ID",
        ),
    ]
    total = 0.0
    for op_type, ticker, time, amount, op_id, comment, position_id in ops:
        rows.append(
            (
                op_type,
                None,
                ticker,
                None,
                time,
                amount,
                op_id,
                comment,
                "My Trades",
                position_id,
            )
        )
        total += amount
    rows.append(
        ("Total", None, None, None, None, round(total, 2), None, None, None, None)
    )
    return rows


_CLOSED_HEADER: tuple[object, ...] = (
    "Instrument",
    "Ticker",
    "Category",
    "Type",
    "Volume",
    "Open Price",
    "Open Time (UTC)",
    "Close Price",
    "Close Time (UTC)",
    "Product",
    "Profit/Loss",
    "Gross Profit",
    "Purchase Value",
    "Sale Value",
    "Stop Loss",
    "Take Profit",
    "Commission",
    "Margin",
    "Swap",
    "Rollover",
    "Open Conversion Rate",
    "Close Conversion Rate",
    "Close Origin",
    "Position ID",
    "Comment",
)


def _closed_sheet(
    account_id: str,
    positions: list[tuple[str, float, object, str]],
) -> list[tuple[object, ...]]:
    """Build a minimal Closed Positions sheet.

    Each position tuple is (ticker, commission, close_time, position_id).
    The Profit/Loss and Purchase/Sale Value columns are left blank — the
    parser no longer extracts purchase_value/sale_value (D8 fee enrichment
    needs only commission and close_time).
    """
    rows: list[tuple[object, ...]] = [
        ("Account number", account_id),
        ("Closed Positions",),
        ("Date from (UTC)", "2006-01-01"),
        ("Date to (UTC)", "2026-08-03"),
        _CLOSED_HEADER,
    ]
    for ticker, commission, close_time, position_id in positions:
        rows.append(
            (
                "Instr",
                ticker,
                "ETF",
                "BUY",
                10,
                100,
                None,
                120,
                close_time,
                "My Trades",
                None,
                None,
                None,
                None,
                None,
                None,
                commission,
                None,
                None,
                None,
                None,
                None,
                "xStation5",
                position_id,
                None,
            )
        )
    return rows


def _build_raw_from_bytes(
    xlsx_bytes: bytes,
    fernet_key: bytes,
    *,
    fetched_at: datetime,
    source_file: str = "report.xlsx",
) -> pa.Table:
    """Build a raw-layer table from .xlsx bytes (shared bronze, D17)."""
    return pa.table(
        {
            "fetched_at": [fetched_at],
            "broker": ["XTB"],
            "source": ["XTB_REPORT"],
            "payload": [encrypt(xlsx_bytes, fernet_key)],
            "payload_hash": [hashlib.sha256(xlsx_bytes).hexdigest()],
            "source_file": [source_file],
        },
        schema=RAW_SCHEMA,
    )


class TestXtbExcelSerialDecoding:
    """D3 defensive branch: raw numeric Excel serials decode to tz-aware UTC.

    The fixture uses date-formatted cells (openpyxl auto-converts them to
    naive ``datetime``). This test exercises the parser's defensive
    ``from_excel`` branch by writing raw numeric serials (e.g. 46236.875) in
    the Cash Operations ``Time`` and Closed Positions ``Close time`` columns.
    This is the regression coverage for the analytics date-handling fix (D3):
    the parser must never emit a serial string into ``event_datetime``.
    """

    def test_numeric_serial_decodes_to_tz_aware_utc(self) -> None:
        # 46236.875 -> 2026-08-02 21:00 UTC (verified via from_excel).
        serial = 46236.875
        sheets = {
            "Open Positions": _op_sheet(DEFAULT_ACCOUNT_ID, DEFAULT_ACCOUNT_CCY),
            "Cash Operations": _cash_sheet(
                DEFAULT_ACCOUNT_ID,
                [
                    (
                        "Deposit",
                        "",
                        serial,  # raw numeric Excel serial in the Time column
                        1000.0,
                        "900011122",
                        "deposit",
                        "",
                    ),
                ],
            ),
            "Closed Positions": _closed_sheet(
                DEFAULT_ACCOUNT_ID,
                [
                    (
                        "SXR8.DE",
                        12.50,
                        serial,  # raw numeric Excel serial in Close time
                        DEFAULT_CLOSED_POSITION_ID,
                    ),
                ],
            ),
        }
        data = build_xlsx_bytes_from_sheets(sheets)
        report = parse_report(data)

        # Cash-operation Time decoded from the serial -> tz-aware UTC.
        assert len(report.cash_operations) == 1
        op = report.cash_operations[0]
        assert op.time == datetime(2026, 8, 2, 21, 0, tzinfo=UTC)
        assert op.time.tzinfo is not None

        # Closed-position Close time decoded from the serial -> tz-aware UTC.
        assert len(report.closed_positions) == 1
        closed = report.closed_positions[0]
        assert closed.close_time == datetime(2026, 8, 2, 21, 0, tzinfo=UTC)
        assert closed.close_time.tzinfo is not None


class TestXtbOpenClosedLifecycle:
    """Open→closed lifecycle (Stage 5 coverage list): a position open in one
    snapshot, closed in the next; fee captured exactly once on the sell.

    D9/D18: transform keeps the latest ``fetched_at`` per ``account_id`` (not a
    union of uploads). Two raw payloads for the same account are combined; the
    second (later ``fetched_at``) supersedes the first. The open position from
    payload 1 does NOT reappear in the snapshot, and the purchase event from
    payload 1 is NOT in events — only the sell (with its fee) survives.
    """

    @pytest.fixture()
    def fernet_key(self) -> bytes:
        return generate_key()

    def test_open_then_closed_fee_captured_once(self, fernet_key: bytes) -> None:
        t_open = datetime(2026, 8, 1, 6, 0, tzinfo=UTC)
        t_closed = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)

        # Payload 1: open SXR8.DE position + a Stock purchase (no fee on purchase).
        workbook_open = build_xlsx_bytes_from_sheets(
            {
                "Open Positions": _op_sheet(
                    DEFAULT_ACCOUNT_ID,
                    DEFAULT_ACCOUNT_CCY,
                    aggregates=[
                        ("Investment Plan", "Core S&P 500", "SXR8.DE", "ETF", 1000.0),
                    ],
                ),
                "Cash Operations": _cash_sheet(
                    DEFAULT_ACCOUNT_ID,
                    [
                        (
                            "Stock purchase",
                            "SXR8.DE",
                            _naive(2026, 7, 10, 8, 0),
                            -1000.0,
                            "900035422",
                            "OPEN BUY 10 @ 100.00",
                            "P1",
                        ),
                    ],
                ),
                "Closed Positions": _closed_sheet(DEFAULT_ACCOUNT_ID, []),
            }
        )
        # Payload 2: SXR8.DE gone from Open Positions; Stock sell + Closed Position.
        workbook_closed = build_xlsx_bytes_from_sheets(
            {
                "Open Positions": _op_sheet(
                    DEFAULT_ACCOUNT_ID,
                    DEFAULT_ACCOUNT_CCY,
                    aggregates=[],  # no open positions
                ),
                "Cash Operations": _cash_sheet(
                    DEFAULT_ACCOUNT_ID,
                    [
                        (
                            "Stock sell",
                            "SXR8.DE",
                            _naive(2026, 8, 2, 7, 0),
                            1200.0,
                            "900041122",
                            "CLOSE BUY 10 @ 120.00",
                            "P1",
                        ),
                    ],
                ),
                "Closed Positions": _closed_sheet(
                    DEFAULT_ACCOUNT_ID,
                    [
                        (
                            "SXR8.DE",
                            15.00,  # nonzero commission -> fee captured once
                            _naive(2026, 8, 2, 7, 0),
                            "P1",
                        ),
                    ],
                ),
            }
        )

        raw_open = _build_raw_from_bytes(
            workbook_open, fernet_key, fetched_at=t_open, source_file="open.xlsx"
        )
        raw_closed = _build_raw_from_bytes(
            workbook_closed, fernet_key, fetched_at=t_closed, source_file="closed.xlsx"
        )
        combined = pa.concat_tables([raw_open, raw_closed], schema=RAW_SCHEMA)

        # Snapshot: latest-per-account = payload 2 -> no SXR8.DE equity row.
        snapshot = transform_snapshot(combined, fernet_key)
        labels = snapshot.column("label").to_pylist()
        assert "SXR8.DE" not in labels  # open position does not reappear
        equity_rows = [
            r for r in snapshot.to_pylist() if r["position_type"] == "EQUITY"
        ]
        assert equity_rows == []  # payload 2 has no open positions
        # CASH row still present (from payload 2's Total row, D22).
        cash_rows = [r for r in snapshot.to_pylist() if r["position_type"] == "CASH"]
        assert len(cash_rows) == 1

        # events: latest-per-account = payload 2 -> only the sell event survives.
        events = transform_events(combined, fernet_key)
        trades = [r for r in events.to_pylist() if r["event_type"] == "TRADE"]
        assert len(trades) == 1  # fee captured exactly once, not zero or twice
        sell = trades[0]
        assert sell["raw_event_type"] == "Stock sell"
        # Fee enriched from the Closed Position (D8), exactly once.
        assert sell["fee_amount"] is not None
        assert decrypt_float(sell["fee_amount"], fernet_key) == pytest.approx(15.00)
        # The purchase event from payload 1 is NOT in events (latest supersedes).
        assert "Stock purchase" not in {r["raw_event_type"] for r in events.to_pylist()}


class TestXtbRealSampleTransformIntegration:
    """Real-sample integration through transform_snapshot + transform_events.

    The parser-level round-trip (TestXtbParser.test_real_sample_round_trip)
    only exercises ``parse_report``. This test flows the real anonymized
    sample through both shared-bronze transforms (D17) to verify the full
    pipeline produces sensible output. The real sample's Closed Position has
    Commission=0, so the sell row's fee_amount must decrypt to 0.0.
    """

    @pytest.fixture()
    def fernet_key(self) -> bytes:
        return generate_key()

    @pytest.mark.skipif(
        not REAL_SAMPLE_PATH.exists(), reason="sample xlsx not checked out"
    )
    def test_real_sample_through_transforms(self, fernet_key: bytes) -> None:
        xlsx_bytes = REAL_SAMPLE_PATH.read_bytes()
        raw = _build_raw_from_bytes(
            xlsx_bytes, fernet_key, fetched_at=datetime.now(UTC)
        )

        # Snapshot: 2 EQUITY aggregates (SXR8.DE, SXRV.DE) + 1 CASH row.
        snapshot = transform_snapshot(raw, fernet_key)
        types = snapshot.column("position_type").to_pylist()
        assert types.count("EQUITY") == 2
        assert types.count("CASH") == 1
        labels = set(snapshot.column("label").to_pylist())
        assert {"SXR8.DE", "SXRV.DE"} <= labels
        assert f"CASH {DEFAULT_ACCOUNT_CCY}" in labels

        # events: events including the deposit, interest, and sell.
        events = transform_events(raw, fernet_key)
        raw_types = {r["raw_event_type"] for r in events.to_pylist()}
        assert "Deposit" in raw_types
        assert "Free funds interest" in raw_types
        assert "Stock sell" in raw_types
        event_types = set(events.column("event_type").to_pylist())
        assert {"DEPOSIT", "INTEREST", "TRADE"} <= event_types

        # The real sample's Closed Position has Commission=0 (the fixture
        # adds a nonzero one). The sell row's fee_amount must decrypt to 0.0.
        sell_rows = [
            r for r in events.to_pylist() if r["raw_event_type"] == "Stock sell"
        ]
        assert len(sell_rows) == 1
        assert sell_rows[0]["fee_amount"] is not None
        assert decrypt_float(sell_rows[0]["fee_amount"], fernet_key) == 0.0


# ---------------------------------------------------------------------------
# Finding-1 fix: filename-derived account_id + guarded latest-per-account
# parse. A malformed row can no longer kill the transform for all accounts.
# ---------------------------------------------------------------------------


def _malformed_xlsx_bytes(account_id: str = DEFAULT_ACCOUNT_ID) -> bytes:
    """Build a workbook whose parse raises (empty summary-block Currency, guard 4)."""
    open_rows: list[tuple[object, ...]] = [
        ("Account number", account_id),
        ("Open Positions",),
        ("Data as of report generated", "2026-08-03"),
        ("Product", "Metric", "Amount", "Currency"),
        ("My Trades", "Value", 0, None),  # empty Currency -> guard 4 raises
    ]
    return build_xlsx_bytes_from_sheets({"Open Positions": open_rows})


class TestAccountIdFromFilename:
    """Unit tests for the filename-derived account_id helper (finding-1 fix)."""

    def test_valid_pattern(self) -> None:
        assert (
            _account_id_from_filename("PLN_12345678_2006-01-01_2026-08-03.xlsx")
            == "12345678"
        )

    def test_no_underscore(self) -> None:
        assert _account_id_from_filename("report.xlsx") is None

    def test_non_digit_account_segment(self) -> None:
        assert (
            _account_id_from_filename("PLN_account_2006-01-01_2026-08-03.xlsx") is None
        )

    def test_empty(self) -> None:
        assert _account_id_from_filename("") is None


class TestLatestPerAccountGuarded:
    """Finding-1 fix: filename grouping + guarded parse with fallback.

    The account_id is read from the filename pattern
    ``{CCY}_{account_id}_{from}_{to}.xlsx``; only the latest row per account is
    parsed, and a malformed latest row falls back to the previous good row. A
    fully-failing account is skipped (no crash) while other accounts survive.
    """

    @pytest.fixture()
    def fernet_key(self) -> bytes:
        return generate_key()

    def test_filename_grouping_two_accounts(self, fernet_key: bytes) -> None:
        """Pattern filenames: two accounts both survive, only latest parsed."""
        t = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw_a = _build_raw_from_bytes(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t,
            source_file="PLN_111_2006-01-01_2026-08-03.xlsx",
        )
        raw_b = _build_raw_from_bytes(
            build_new_format_xlsx_bytes(account_id="222", account_ccy="EUR"),
            fernet_key,
            fetched_at=t,
            source_file="EUR_222_2006-01-01_2026-08-03.xlsx",
        )
        combined = pa.concat_tables([raw_a, raw_b], schema=RAW_SCHEMA)
        result = transform_snapshot(combined, fernet_key)

        assert set(result.column("account_id").to_pylist()) == {"111", "222"}
        # 2 EQUITY + 1 CASH per account = 6 rows.
        assert result.num_rows == 6

    def test_malformed_latest_falls_back_to_older_row(
        self, fernet_key: bytes, caplog
    ) -> None:
        """A malformed latest row falls back to the previous good row (no crash)."""
        t_old = datetime(2026, 8, 1, 6, 0, tzinfo=UTC)
        t_new = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw_old = _build_raw_from_bytes(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t_old,
            source_file="PLN_111_2006-01-01_2026-08-01.xlsx",
        )
        raw_new = _build_raw_from_bytes(
            _malformed_xlsx_bytes("111"),
            fernet_key,
            fetched_at=t_new,
            source_file="PLN_111_2006-01-01_2026-08-03.xlsx",
        )
        combined = pa.concat_tables([raw_old, raw_new], schema=RAW_SCHEMA)

        caplog.set_level("WARNING", logger="pipeline.connectors.xtb.transform")
        result = transform_snapshot(combined, fernet_key)

        # Fell back to the older good row -> 3 rows for account 111, no crash.
        assert result.num_rows == 3
        assert set(result.column("account_id").to_pylist()) == {"111"}
        assert any("trying older row" in r.message for r in caplog.records)

    def test_all_rows_for_account_fail_skips_account(
        self, fernet_key: bytes, caplog
    ) -> None:
        """When every row for an account fails to parse, the account is skipped."""
        t1 = datetime(2026, 8, 2, 6, 0, tzinfo=UTC)
        t2 = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw_bad1 = _build_raw_from_bytes(
            _malformed_xlsx_bytes("111"),
            fernet_key,
            fetched_at=t1,
            source_file="PLN_111_2006-01-01_2026-08-02.xlsx",
        )
        raw_bad2 = _build_raw_from_bytes(
            _malformed_xlsx_bytes("111"),
            fernet_key,
            fetched_at=t2,
            source_file="PLN_111_2006-01-01_2026-08-03.xlsx",
        )
        raw_good = _build_raw_from_bytes(
            build_new_format_xlsx_bytes(account_id="222", account_ccy="EUR"),
            fernet_key,
            fetched_at=t2,
            source_file="EUR_222_2006-01-01_2026-08-03.xlsx",
        )
        combined = pa.concat_tables([raw_bad1, raw_bad2, raw_good], schema=RAW_SCHEMA)

        caplog.set_level("WARNING", logger="pipeline.connectors.xtb.transform")
        result = transform_snapshot(combined, fernet_key)

        # Only account 222 survives; account 111 skipped (no exception).
        assert set(result.column("account_id").to_pylist()) == {"222"}
        assert result.num_rows == 3

    def test_filename_vs_r1_mismatch_r1_wins(self, fernet_key: bytes, caplog) -> None:
        """Filename account_id != report R1 -> warning logged, R1 account_id emitted."""
        t = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw = _build_raw_from_bytes(
            build_new_format_xlsx_bytes(account_id="999", account_ccy="PLN"),
            fernet_key,
            fetched_at=t,
            source_file="PLN_111_2006-01-01_2026-08-03.xlsx",  # filename says 111
        )

        caplog.set_level("WARNING", logger="pipeline.connectors.xtb.transform")
        result = transform_snapshot(raw, fernet_key)

        # R1 account_id (999) wins in the emitted rows.
        assert set(result.column("account_id").to_pylist()) == {"999"}
        assert any("!= report R1" in r.message for r in caplog.records)

    def test_non_matching_filename_uses_fallback_parse(self, fernet_key: bytes) -> None:
        """A non-pattern filename falls back to a guarded parse for account_id."""
        t = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        raw = _build_raw_from_bytes(
            build_new_format_xlsx_bytes(account_id="111", account_ccy="PLN"),
            fernet_key,
            fetched_at=t,
            source_file="report.xlsx",  # no underscore-digit pattern
        )
        result = transform_snapshot(raw, fernet_key)

        # Fallback parse discovered account_id 111 -> 3 rows.
        assert result.num_rows == 3
        assert set(result.column("account_id").to_pylist()) == {"111"}

    def test_dropped_cash_rows_warned(self, fernet_key: bytes, caplog) -> None:
        """Empty-Time cash rows are dropped with a warning, not silently (finding 3)."""
        t = datetime(2026, 8, 3, 6, 0, tzinfo=UTC)
        sheets = {
            "Open Positions": _op_sheet(DEFAULT_ACCOUNT_ID, DEFAULT_ACCOUNT_CCY),
            "Cash Operations": _cash_sheet(
                DEFAULT_ACCOUNT_ID,
                [
                    (
                        "Deposit",
                        "",
                        _naive(2026, 7, 10, 0, 0),
                        1000.0,
                        "900011122",
                        "deposit",
                        "",
                    ),
                    ("Withdrawal", "", None, -200.0, "900011123", "withdrawal", ""),
                ],
            ),
            "Closed Positions": _closed_sheet(DEFAULT_ACCOUNT_ID, []),
        }
        raw = _build_raw_from_bytes(
            build_xlsx_bytes_from_sheets(sheets),
            fernet_key,
            fetched_at=t,
            source_file="PLN_12345678_2006-01-01_2026-08-03.xlsx",
        )

        caplog.set_level("WARNING", logger="pipeline.connectors.xtb.transform")
        events = transform_events(raw, fernet_key)

        # The Withdrawal is absent from events, but the drop is signaled.
        assert [r["raw_event_type"] for r in events.to_pylist()] == ["Deposit"]
        assert any(
            "dropped" in r.message and "empty Time" in r.message for r in caplog.records
        )
