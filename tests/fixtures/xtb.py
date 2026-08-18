"""XTB fixture builders for the new-format 3-sheet workbook and Delta tables.

The new XTB export format (D13) is a 3-sheet .xlsx: "Open Positions",
"Cash Operations", "Closed Positions". `build_new_format_xlsx_bytes`
constructs such a workbook programmatically with openpyxl using known,
assertable values (including a Closed Position with a NONZERO commission,
which the real anonymized sample lacks). The real sample file at
``docs/xtb/xtb-report-sample/PLN_12345678_2006-01-01_2026-08-03.xlsx`` is kept
as a secondary integration fixture.

`xtb_raw_snapshot` / `xtb_normalized_snapshot` build raw/normalized Delta
tables for other connector test files; they are kept compatible so those
test files continue to import cleanly.
"""

from __future__ import annotations

import hashlib
from datetime import UTC, datetime
from io import BytesIO

import openpyxl
import pyarrow as pa
from openpyxl.workbook.workbook import Workbook

from pipeline.crypto import encrypt, encrypt_float, generate_key
from pipeline.normalized.models import snapshot_normalized_schema
from pipeline.raw.models import RAW_SCHEMA


def _dt(year: int, month: int, day: int, hour: int = 0, minute: int = 0) -> datetime:
    """Build a naive datetime for an XLSX date cell.

    openpyxl rejects tz-aware datetimes on write ("Excel does not support
    timezones"), and the real XTB export stores naive date cells that the
    parser attaches UTC to at the read boundary (D3). The single ``# noqa:
    DTZ001`` here suppresses the ruff naive-datetime warning for the whole
    module so callers can build date cells without per-line annotations.
    """
    return datetime(year, month, day, hour, minute)  # noqa: DTZ001


# ---------------------------------------------------------------------------
# Default known values for the programmatic fixture. Tests assert against
# these (or against the overridden parameters passed to the builder).
# ---------------------------------------------------------------------------

DEFAULT_ACCOUNT_ID = "12345678"
DEFAULT_ACCOUNT_CCY = "PLN"
DEFAULT_CLOSED_POSITION_ID = "1111122222"
DEFAULT_CLOSED_COMMISSION = 12.50
DEFAULT_CLOSED_PURCHASE_VALUE = 4300.04
DEFAULT_CLOSED_SALE_VALUE = 5040.05

# Cash-operation known values (kept as module constants for test assertions).
DEPOSIT_AMOUNT = 10000.00
INTEREST_AMOUNT = 100.01
INTEREST_TAX_AMOUNT = -19.00
SUBACCOUNT_IN_AMOUNT = 10000.00
SUBACCOUNT_OUT_AMOUNT = -10000.00
CONVERSION_TRANSFER_AMOUNT = -1000.00
PURCHASE_AMOUNT = -4300.04
SELL_AMOUNT = 5040.05

# Open-position aggregate known values.
SXR8_AGGREGATE_VALUE = 5544.00
SXRV_AGGREGATE_VALUE = 3633.84
SOLD_OUT_AGGREGATE_VALUE = 0.00  # zero-value aggregate -> guard 3 skip


def _fill_workbook(
    sheets: dict[str, list[tuple[object, ...]]],
) -> Workbook:
    """Build an openpyxl Workbook from a mapping of sheet name -> list of rows."""
    wb = openpyxl.Workbook()
    # Remove the default sheet; we create all sheets explicitly.
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row_idx, row_values in enumerate(rows, start=1):
            for col_idx, value in enumerate(row_values, start=1):
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)
    return wb


def build_xlsx_bytes_from_sheets(sheets: dict[str, list[tuple[object, ...]]]) -> bytes:
    """Build an .xlsx from an arbitrary mapping of sheet name -> rows.

    Used by parser edge-case tests (e.g. a missing sheet, or an Open
    Positions sheet with an empty summary-block Currency) that need a
    shape the default ``build_new_format_xlsx_bytes`` does not produce.
    """
    wb = _fill_workbook(sheets)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_new_format_xlsx_bytes(
    account_id: str = DEFAULT_ACCOUNT_ID,
    account_ccy: str = DEFAULT_ACCOUNT_CCY,
    *,
    closed_position_id: str = DEFAULT_CLOSED_POSITION_ID,
    closed_commission: float = DEFAULT_CLOSED_COMMISSION,
    closed_purchase_value: float = DEFAULT_CLOSED_PURCHASE_VALUE,
    closed_sale_value: float = DEFAULT_CLOSED_SALE_VALUE,
) -> bytes:
    """Build a new-format XTB report workbook as .xlsx bytes.

    The workbook has the 3 sheets "Open Positions", "Cash Operations",
    "Closed Positions" matching the real new-format layout (R1 Account
    number, R2 sheet title, R3/R4 date range, then header + data rows).

    Known values (assertable in tests):
      - account_id / account_ccy from the Open Positions summary block.
      - Open Positions: 2 instruments (SXR8.DE, SXRV.DE) each with an
        aggregate row + child lot row, plus a zero-value aggregate
        (SOLD.DE) to exercise guard 3.
      - Cash Operations: Deposit, Free funds interest, Free funds interest
        tax, a Subaccount transfer +/- pair (net zero, filtered by D7), a
        currency-conversion Transfer, a Stock purchase, a Stock sell (its
        Position ID matches the Closed Position), and a Total row whose
        Amount equals the sum of the other amounts.
      - Closed Positions: one row with a NONZERO commission matching the
        Stock sell's Position ID, plus a Profit/loss total row.
    """
    sheets = {
        "Open Positions": _open_positions_rows(account_id, account_ccy),
        "Cash Operations": _cash_operations_rows(account_id),
        "Closed Positions": _closed_positions_rows(
            account_id,
            closed_position_id,
            closed_commission,
            closed_purchase_value,
            closed_sale_value,
        ),
    }
    wb = _fill_workbook(sheets)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _open_positions_rows(account_id: str, account_ccy: str) -> list[tuple[object, ...]]:
    """Build the Open Positions sheet rows (summary block + detail rows)."""
    # R1-R3: account header + report title + generation timestamp.
    rows: list[tuple[object, ...]] = [
        ("Account number", account_id),
        ("Open Positions",),
        ("Data as of report generated", _dt(2026, 8, 3, 6, 0)),
        # R4: summary block header. The Currency column (col 4) carries the
        # account currency (D5); the parser reads it then skips this block.
        ("Product", "Metric", "Amount", "Currency"),
        ("My Trades", "Value", 0, account_ccy),
        ("My Trades", "Profit", 0, account_ccy),
        (
            "Investment Plans",
            "Value",
            SXR8_AGGREGATE_VALUE + SXRV_AGGREGATE_VALUE,
            account_ccy,
        ),
        ("Investment Plans", "Profit", 1000.00, account_ccy),
        # R9: empty row (matches the real sample layout).
        (None,),
        # R10: note row.
        (
            "Note",
            "Summary values and open positions are shown as of the report generation time",
        ),
        # R11: detail header.
        (
            "Product",
            "Instrument/Position",
            "Ticker",
            "Category",
            "Type",
            "Volume",
            "Value",
            "Current price",
            "Open price",
            "Open time (UTC)",
            "Stop Loss",
            "Take Profit",
            "Net Profit %",
            "Net Profit",
            "Gross Profit",
            "Margin",
            "Open Commission",
            "Swap",
            "Rollover",
        ),
    ]

    # R12: aggregate row for SXR8.DE (empty Type, real name in Instrument,
    # non-empty Category) -> kept as a holding.
    rows += [
        (
            "Investment Plan",
            "Core S&P 500",
            "SXR8.DE",
            "ETF",
            None,
            11,
            SXR8_AGGREGATE_VALUE,
            None,
            106.36,
            None,
            None,
            None,
            29.77,
            626.90,
            626.90,
            None,
            None,
            None,
        ),
        # R13: child lot for SXR8.DE (numeric ID in Instrument, empty
        # Category, non-empty Type) -> skipped by the parser (D4).
        (
            "Investment Plan",
            "1334567890",
            "SXR8.DE",
            None,
            "BUY",
            7,
            3528.00,
            120,
            110,
            _dt(2026, 8, 2, 8, 0),
            None,
            None,
            8.32,
            270.90,
            270.90,
        ),
        # R14: aggregate row for SXRV.DE.
        (
            "Investment Plan",
            "NASDAQ 100",
            "SXRV.DE",
            "ETF",
            None,
            4,
            SXRV_AGGREGATE_VALUE,
            None,
            200,
            None,
            None,
            None,
            9.45,
            313.84,
            313.84,
            None,
            None,
            None,
        ),
        # R15: child lot for SXRV.DE -> skipped.
        (
            "Investment Plan",
            "1244567890",
            "SXRV.DE",
            None,
            "BUY",
            4,
            SXRV_AGGREGATE_VALUE,
            216.30,
            200,
            _dt(2026, 7, 10, 8, 0),
            None,
            None,
            9.45,
            313.84,
            313.84,
            None,
            None,
            None,
        ),
        # R16: zero-value aggregate for SOLD.DE -> skipped by guard 3.
        (
            "Investment Plan",
            "Sold Out Inc",
            "SOLD.DE",
            "ETF",
            None,
            0,
            SOLD_OUT_AGGREGATE_VALUE,
            None,
            None,
            None,
            None,
            None,
            0,
            0,
            0,
            None,
            None,
            None,
        ),
    ]
    return rows


def _cash_operations_rows(account_id: str) -> list[tuple[object, ...]]:
    """Build the Cash Operations sheet rows."""
    # The Total Amount equals the sum of the other amounts (subaccount
    # transfers net to zero, so including or excluding them is the same).
    total = (
        DEPOSIT_AMOUNT
        + INTEREST_AMOUNT
        + INTEREST_TAX_AMOUNT
        + SUBACCOUNT_IN_AMOUNT
        + SUBACCOUNT_OUT_AMOUNT
        + CONVERSION_TRANSFER_AMOUNT
        + PURCHASE_AMOUNT
        + SELL_AMOUNT
    )
    # R1-R4: account header + title + date range.
    rows: list[tuple[object, ...]] = [
        ("Account number", account_id),
        ("Cash Operations",),
        ("Date from (UTC)", _dt(2006, 1, 1, 0, 0)),
        ("Date to (UTC)", _dt(2026, 8, 3, 6, 0)),
        # R5: header.
        (
            "Type",
            "Instrument",
            "Ticker",
            "Category",
            "Time",
            "Amount",
            "ID",
            "Comment",
            "Product",
            "Position ID",
        ),
        # R6: Deposit.
        (
            "Deposit",
            None,
            None,
            None,
            _dt(2026, 7, 10, 0, 0),
            DEPOSIT_AMOUNT,
            900011122,
            "Adyen BLIK deposit, id=22979000",
            "My Trades",
            None,
        ),
        # R7: Free funds interest.
        (
            "Free funds interest",
            None,
            None,
            None,
            _dt(2026, 8, 2, 21, 0),
            INTEREST_AMOUNT,
            1000011122,
            "Free-funds Interest 2026-07",
            "My Trades",
            None,
        ),
        # R8: Free funds interest tax.
        (
            "Free funds interest tax",
            None,
            None,
            None,
            _dt(2026, 8, 2, 21, 1),
            INTEREST_TAX_AMOUNT,
            1000011142,
            "Free-funds Interest Tax 2026-07",
            "My Trades",
            None,
        ),
        # R9: Subaccount transfer IN -> filtered by D7.
        (
            "Subaccount transfer",
            None,
            None,
            None,
            _dt(2026, 7, 10, 6, 30),
            SUBACCOUNT_IN_AMOUNT,
            900011174,
            "Transfer from 12345678 to 12348765",
            "Investment Plans",
            None,
        ),
        # R10: Subaccount transfer OUT -> filtered by D7.
        (
            "Subaccount transfer",
            None,
            None,
            None,
            _dt(2026, 7, 10, 6, 30),
            SUBACCOUNT_OUT_AMOUNT,
            900011173,
            "Transfer from 12345678 to 12348765",
            "My Trades",
            None,
        ),
        # R11: currency-conversion Transfer (kept; target_fx_rate null, D7).
        (
            "Transfer",
            None,
            None,
            None,
            _dt(2026, 8, 2, 20, 0),
            CONVERSION_TRANSFER_AMOUNT,
            900051122,
            "Currency conversion, PLN to EUR from TA: 12345678 to: 12345670, Exchange rate:0.230001",
            "My Trades",
            None,
        ),
        # R12: Stock purchase (position_id different from the sell).
        (
            "Stock purchase",
            "Core S&P 500",
            "SXR8.DE",
            "ETF",
            _dt(2026, 7, 10, 8, 0),
            PURCHASE_AMOUNT,
            900035422,
            "OPEN BUY 10.0001 @ 100.00",
            "Investment Plans",
            1244567890,
        ),
        # R13: Stock sell (Position ID matches the Closed Position row).
        (
            "Stock sell",
            "Core S&P 500",
            "SXR8.DE",
            "ETF",
            _dt(2026, 8, 2, 7, 0),
            SELL_AMOUNT,
            900041122,
            "CLOSE BUY 10.0001 @ 120.00",
            "Investment Plans",
            DEFAULT_CLOSED_POSITION_ID,
        ),
        # R14: Total row -> read into free_cash, then excluded from events.
        (
            "Total",
            None,
            None,
            None,
            None,
            round(total, 2),
            None,
            None,
            None,
            None,
        ),
    ]
    return rows


def _closed_positions_rows(
    account_id: str,
    position_id: str,
    commission: float,
    purchase_value: float,
    sale_value: float,
) -> list[tuple[object, ...]]:
    """Build the Closed Positions sheet rows."""
    rows: list[tuple[object, ...]] = [
        ("Account number", account_id),
        ("Closed Positions",),
        ("Date from (UTC)", _dt(2006, 1, 1, 0, 0)),
        ("Date to (UTC)", _dt(2026, 8, 3, 6, 0)),
        # R5: header (full column set per the real sample).
        (
            "Instrument",
            "Ticker",
            "Category",
            "Type",
            "Volume",
            "Open Price",
            "Open Time (UTC)",
            "Close Price",
            "Close Time (UTC)",
            "Product",
            "Profit/Loss",
            "Gross Profit",
            "Purchase Value",
            "Sale Value",
            "Stop Loss",
            "Take Profit",
            "Commission",
            "Margin",
            "Swap",
            "Rollover",
            "Open Conversion Rate",
            "Close Conversion Rate",
            "Close Origin",
            "Position ID",
            "Comment",
        ),
        # R6: one closed trade with a NONZERO commission.
        (
            "Core S&P 500",
            "SXR8.DE",
            "ETF",
            "BUY",
            10.0001,
            100,
            _dt(2026, 7, 1, 7, 0),
            120,
            _dt(2026, 8, 2, 7, 0),
            "Investment Plans",
            sale_value - purchase_value,
            sale_value - purchase_value,
            purchase_value,
            sale_value,
            None,
            None,
            commission,
            None,
            None,
            None,
            4.3,
            4.2,
            "xStation5",
            position_id,
            None,
        ),
        # R7: Profit/loss total row -> excluded by the parser.
        (
            "Profit/loss",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            sale_value - purchase_value,
            sale_value - purchase_value,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ),
    ]
    return rows


# ---------------------------------------------------------------------------
# Delta-table fixture helpers (kept compatible for other test files).
# ---------------------------------------------------------------------------


def xtb_raw_snapshot(
    fernet_key: bytes | None = None,
) -> pa.Table:
    """Build a raw XTB snapshot table with an encrypted new-format workbook.

    The payload is a new-format 3-sheet .xlsx (binary). The ``source`` is
    ``"XTB_REPORT"`` per D17 (shared bronze: one raw row carries all 3
    sheets and feeds both snapshot and CDC silvers).
    """
    if fernet_key is None:
        fernet_key = generate_key()

    now = datetime.now(UTC)
    payload = build_new_format_xlsx_bytes()
    encrypted_payload = encrypt(payload, fernet_key)

    return pa.table(
        {
            "fetched_at": [now],
            "broker": ["xtb"],
            "source": ["XTB_REPORT"],
            "payload": [encrypted_payload],
            "payload_hash": [hashlib.sha256(payload).hexdigest()],
            "source_file": ["report.xlsx"],
        },
        schema=RAW_SCHEMA,
    )


def xtb_normalized_snapshot(
    fernet_key: bytes | None = None,
    account_id: str = "XTB-12345",
) -> pa.Table:
    """Build a normalized XTB snapshot table with encrypted values.

    Default data: 2 equities (VWCE.DE, CDR.PL) + 1 cash entry (PLN).
    Kept as a compatible stub for the test files that import it
    (transform/consolidate/report/portfolio fixtures); the parser/transform
    rewrite does not change this helper's schema.
    """
    if fernet_key is None:
        fernet_key = generate_key()
    now = datetime.now(UTC)
    return pa.table(
        {
            "fetched_at": [now, now, now],
            "account_id": [account_id, account_id, account_id],
            "position_type": ["EQUITY", "EQUITY", "CASH"],
            "label": ["VWCE.DE", "CDR.PL", "CASH:PLN"],
            "description": [
                "Vanguard FTSE All-World UCITS ETF",
                "CD Projekt",
                "Cash PLN",
            ],
            "asset_class": ["STK", "STK", "CASH"],
            "security_value": [
                encrypt_float(1000.0, fernet_key),
                encrypt_float(2500.0, fernet_key),
                encrypt_float(5000.0, fernet_key),
            ],
            "security_ccy": ["EUR", "PLN", "PLN"],
            "isin": ["IE00BK5BQT80", "PL9999900006", ""],
        },
        schema=snapshot_normalized_schema,
    )
